import io
import re
from pathlib import Path

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="ASN PDF Carton Calculator", layout="wide")

ASN_RE = re.compile(r"ASN\s*No\s*:\s*((?:CH|CR)\d+)", re.IGNORECASE)


def clean(v):
    return "" if v is None else str(v).replace("\r", "").strip()


def safe_int(v):
    s = clean(v).replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def norm_rev_output(v):
    s = clean(v)
    if not s:
        return ""
    try:
        return f"Rev{int(float(s)):02d}"
    except Exception:
        return s


@st.cache_data(show_spinner=False)
def load_packing_db_from_bytes(xlsx_bytes: bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    if "LOOKUP_TABLE" in wb.sheetnames:
        ws = wb["LOOKUP_TABLE"]
    elif "MASTER_DB" in wb.sheetnames:
        ws = wb["MASTER_DB"]
    elif "Packing_DB" in wb.sheetnames:
        ws = wb["Packing_DB"]
    else:
        ws = wb[wb.sheetnames[0]]

    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = clean(row[0] if len(row) > 0 else "")
        pcs = safe_int(row[2] if len(row) > 2 else (row[1] if len(row) > 1 else ""))
        if item and pcs:
            lookup[item] = pcs
    return lookup


def detect_asn(text, fallback_name):
    m = ASN_RE.search(text or "")
    return m.group(1) if m else fallback_name


def table_has_required_headers(header_row):
    hdr = [clean(c).lower().replace("\n", " ") for c in header_row]
    return ("item no." in hdr) and ("quantity" in hdr)


def extract_rows_from_pdf(uploaded_file, packing_lookup):
    pdf_bytes = uploaded_file.read()
    pdf_name = uploaded_file.name
    all_rows = []
    asn = ""

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            if not asn:
                asn = detect_asn(text, Path(pdf_name).stem)

            for table in page.extract_tables() or []:
                if not table or not table[0]:
                    continue
                if not table_has_required_headers(table[0]):
                    continue

                for raw in table[1:]:
                    if not raw:
                        continue

                    row = [clean(c) for c in raw]
                    first = clean(row[0]).lower()

                    if first.startswith("total quantity"):
                        continue
                    if not clean(row[0]).isdigit():
                        continue

                    item = clean(row[2] if len(row) > 2 else "")
                    rev = norm_rev_output(row[3] if len(row) > 3 else "")
                    qty = safe_int(row[4] if len(row) > 4 else "") or 0
                    uom = clean(row[5] if len(row) > 5 else "")
                    net_weight = clean(row[6] if len(row) > 6 else "")
                    lot_invoice = clean(row[9] if len(row) > 9 else "").replace("\n", " | ")
                    line_no = clean(row[10] if len(row) > 10 else "")
                    pcs_ctn = packing_lookup.get(item, "")

                    all_rows.append({
                        "ASN": asn or Path(pdf_name).stem,
                        "Seq": safe_int(row[0]),
                        "PO No": clean(row[1] if len(row) > 1 else ""),
                        "Item No": item,
                        "Rev": rev,
                        "Quantity": qty,
                        "Uom": uom,
                        "Net Weight (KG)": net_weight,
                        "Line No": line_no,
                        "Lot/Invoice": lot_invoice,
                        "PCS/CTN": pcs_ctn,
                        "Source PDF": pdf_name,
                        "Page": page_num,
                    })

    return all_rows


def group_data(raw_df):
    if raw_df.empty:
        return pd.DataFrame(columns=[
            "ASN", "Item No", "Rev", "Line No", "Total Qty",
            "PCS/CTN", "Cartons", "Loose PCS", "Status"
        ])

    grouped = (
        raw_df.groupby(["ASN", "Item No", "Rev", "Line No"], dropna=False, as_index=False)
        .agg({"Quantity": "sum", "PCS/CTN": "first"})
        .rename(columns={"Quantity": "Total Qty"})
    )

    grouped["Cartons"] = grouped.apply(
        lambda r: int(r["Total Qty"]) // int(r["PCS/CTN"]) if str(r["PCS/CTN"]).strip() != "" else "",
        axis=1
    )
    grouped["Loose PCS"] = grouped.apply(
        lambda r: int(r["Total Qty"]) % int(r["PCS/CTN"]) if str(r["PCS/CTN"]).strip() != "" else "",
        axis=1
    )
    grouped["Status"] = grouped["PCS/CTN"].apply(lambda x: "OK" if str(x).strip() != "" else "NO PACKING DB")

    return grouped[[
        "ASN", "Item No", "Rev", "Line No", "Total Qty",
        "PCS/CTN", "Cartons", "Loose PCS", "Status"
    ]]


def build_summary(grouped_df, raw_df):
    if grouped_df.empty:
        return pd.DataFrame(columns=[
            "ASN", "Lines", "Grouped Rows", "Total Qty",
            "Total Cartons", "Total Loose PCS", "Missing Packing Rows"
        ])

    lines_per_asn = raw_df.groupby("ASN", as_index=False).size().rename(columns={"size": "Lines"})
    grouped_per_asn = grouped_df.groupby("ASN", as_index=False).size().rename(columns={"size": "Grouped Rows"})
    qty_per_asn = grouped_df.groupby("ASN", as_index=False)["Total Qty"].sum()
    carton_per_asn = grouped_df.groupby("ASN", as_index=False)["Cartons"].apply(
        lambda s: sum(int(x) for x in s if str(x).strip() != "")
    ).rename(columns={"Cartons": "Total Cartons"})
    loose_per_asn = grouped_df.groupby("ASN", as_index=False)["Loose PCS"].apply(
        lambda s: sum(int(x) for x in s if str(x).strip() != "")
    ).rename(columns={"Loose PCS": "Total Loose PCS"})
    miss_per_asn = grouped_df.groupby("ASN", as_index=False)["Status"].apply(
        lambda s: sum(1 for x in s if x != "OK")
    ).rename(columns={"Status": "Missing Packing Rows"})

    summary = lines_per_asn
    for df in [grouped_per_asn, qty_per_asn, carton_per_asn, loose_per_asn, miss_per_asn]:
        summary = summary.merge(df, on="ASN", how="left")
    return summary


def build_excel(raw_df, grouped_df, summary_df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        raw_df.to_excel(writer, index=False, sheet_name="ASN_Raw_Data")
        grouped_df.to_excel(writer, index=False, sheet_name="ASN_Grouped")
        summary_df.to_excel(writer, index=False, sheet_name="ASN_Summary")
    output.seek(0)
    return output


st.title("ASN PDF Carton Calculator")
st.caption("Lookup packing theo Item No. Gop theo ASN + Item No + Line No. Van giu Rev o dau ra.")

col1, col2 = st.columns(2)

with col1:
    pdf_files = st.file_uploader("Upload Delivery Note PDF", type=["pdf"], accept_multiple_files=True)

with col2:
    packing_db = st.file_uploader("Upload Packing DB", type=["xlsx"])

run = st.button("Run", use_container_width=True)

if run:
    if not pdf_files:
        st.error("Ban chua upload PDF.")
        st.stop()
    if not packing_db:
        st.error("Ban chua upload Packing DB.")
        st.stop()

    with st.spinner("Dang doc Packing DB..."):
        packing_lookup = load_packing_db_from_bytes(packing_db.read())

    if not packing_lookup:
        st.error("Khong doc duoc Packing DB.")
        st.stop()

    raw_rows = []
    with st.spinner("Dang doc PDF..."):
        for pdf in pdf_files:
            pdf.seek(0)
            raw_rows.extend(extract_rows_from_pdf(pdf, packing_lookup))

    raw_df = pd.DataFrame(raw_rows)
    grouped_df = group_data(raw_df)
    summary_df = build_summary(grouped_df, raw_df)

    st.success(f"Da xu ly xong {len(pdf_files)} file PDF.")

    st.subheader("ASN Summary")
    st.dataframe(summary_df, use_container_width=True)

    st.subheader("ASN Grouped")
    st.dataframe(grouped_df, use_container_width=True)

    st.subheader("ASN Raw Data")
    st.dataframe(raw_df, use_container_width=True)

    excel_file = build_excel(raw_df, grouped_df, summary_df)
    st.download_button(
        "Download ASN_Result.xlsx",
        data=excel_file,
        file_name="ASN_Result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
